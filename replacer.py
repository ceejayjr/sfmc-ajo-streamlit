import re
from typing import List, Tuple, Dict
from io import BytesIO
from openpyxl import load_workbook

def _load_mapping_from_xlsx(excel_bytes: bytes) -> List[Tuple[str, str]]:
    """
    Lê o Excel em memória (A=SFMC, B=AJO; cabeçalhos na linha 1)
    Retorna lista de pares (sfmc, ajo) como strings.
    """
    wb = load_workbook(filename=BytesIO(excel_bytes), read_only=True, data_only=True)
    ws = wb.active  # primeira aba
    rows = list(ws.iter_rows(min_row=2, max_col=2, values_only=True))  # a partir da linha 2

    mapping = []
    for sfmc, ajo in rows:
        if sfmc is None:
            continue
        sfmc_str = str(sfmc).strip()
        ajo_str = "" if ajo is None else str(ajo).strip()
        if sfmc_str:
            mapping.append((sfmc_str, ajo_str))
    return mapping

def _build_flex_regex(sfmc_snippet: str) -> re.Pattern:
    """
    Regex 'flexível':
      - ignora variações de espaço/linha -> \s*
      - aceita espaços opcionais ao redor de '='
      - funciona com tokens colados (%%...%%)
    """
    s = sfmc_snippet.strip()
    esc = re.escape(s)
    # qualquer espaço/quebra vira \s*
    esc = re.sub(r'\\[ \t\r\n\f]+', r'\\s*', esc)
    # '=' com espaços opcionais
    esc = esc.replace(r'\=', r'\s*=\s*')
    return re.compile(esc, flags=re.IGNORECASE | re.DOTALL)

def _comment_ampscript(html_text: str) -> Tuple[str, int]:
    """
    Comenta AMPScript remanescente:
      - %%= ... =%%
      - %%[ ... ]%%
      - <script runat="server">...</script>
    Não comenta blocos AJO ({{ }} e {% %}).
    """
    patterns = [
        r'%%=.+?=%%',
        r'%%\[[\s\S]*?\]%%',
        r'<script[^>]*\brunat=[\'"]server[\'"][^>]*>[\s\S]*?<\/script>',
    ]
    combined = re.compile("|".join(f"(?:{p})" for p in patterns),
                          flags=re.IGNORECASE | re.DOTALL)

    count = 0
    def _wrap(m):
        nonlocal count
        txt = m.group(0)
        st = txt.strip()
        if st.startswith("<!--") and st.endswith("-->"):
            return txt
        if ("{{" in txt and "}}" in txt) or ("{%" in txt and "%}" in txt):
            return txt
        count += 1
        return f"<!-- {txt} -->"

    new_html = combined.sub(_wrap, html_text)
    return new_html, count

def process(html_bytes: bytes, excel_bytes: bytes):
    """
    Recebe bytes do HTML e XLSX, retorna (html_processado_bytes, relatório_dict)
    """
    mapping = _load_mapping_from_xlsx(excel_bytes)

    # Ordena por SFMC mais longo primeiro (evita match parcial antes do completo)
    mapping.sort(key=lambda p: len(p[0].strip()), reverse=True)

    html = html_bytes.decode("utf-8", errors="replace").replace("\r\n", "\n")

    total_found = 0
    total_replaced = 0
    subs_log: List[Tuple[str, int]] = []

    for sfmc, ajo in mapping:
        if not sfmc:
            continue
        pattern = _build_flex_regex(sfmc)
        matches = list(pattern.finditer(html))
        if not matches:
            continue
        total_found += len(matches)
        if ajo:
            html, n = pattern.subn(lambda _m, rep=ajo: rep, html)
            total_replaced += n
            subs_log.append((sfmc[:80] + ("..." if len(sfmc) > 80 else ""), n))

    html, commented = _comment_ampscript(html)

    report = {
        "found": total_found,
        "replaced": total_replaced,
        "commented": commented,
        "details": subs_log[:50],
    }
    return html.encode("utf-8"), report
